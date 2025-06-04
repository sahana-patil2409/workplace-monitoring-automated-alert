import cv2
import smtplib
import time
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
import numpy as np
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pyttsx3
import winsound
engine = pyttsx3.init()
empthreshold=60
# Email settings
sender_email = "reethugowda07@gmail.com"
receiver_email = "workplacemonitoringfyp@googlegroups.com"
email_password = " abcdutxqyz"  # Use an app-specific password if using Gmail
smtp_server = "smtp.gmail.com"
smtp_port = 587 

# Load YOLO
net = cv2.dnn.readNet("person.weights", "person.cfg")
layer_names = net.getLayerNames()
output_layers = [layer_names[i - 1] for i in net.getUnconnectedOutLayers()]

# Load the class names
with open("coco.names", "r") as f:
    classes = [line.strip() for line in f.readlines()]

# Function to send email with image attachment
def send_email(image_path, subject, body):
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

    # Attach image
    with open(image_path, 'rb') as img_file:
        img = MIMEImage(img_file.read())
        msg.attach(img)

    # Send email
    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, email_password)
        server.sendmail(sender_email, receiver_email, msg.as_string())
        server.quit()
        print("Email sent successfully")
    except Exception as e:
        print(f"Failed to send email: {e}")

# Initialize variables
human_detected = False
last_detection_time = time.time()
person_detected_frame = None
detection_label = "Employee Not Detected"
email_sent = False
log_filename = "employee_status.xlsx"
notdetectedtime=None
# Initialize the Excel workbook
try:
    workbook = load_workbook(log_filename)
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Date", "Status", "Time"])

# Open the video capture
cap = cv2.VideoCapture(0)  # Use the first camera

def log_status(status, time):
    sheet.append([datetime.now().strftime('%Y-%m-%d %H:%M:%S'), status, time])
    workbook.save(log_filename)

while True:
    ret, frame = cap.read()
    if not ret:
        print("Failed to capture frame")
        break

    height, width, channels = frame.shape
    blob = cv2.dnn.blobFromImage(frame, 0.00392, (608, 608), (0, 0, 0), True, crop=False)
    net.setInput(blob)
    outs = net.forward(output_layers)

    # Variables for bounding box and person detection
    boxes = []
    confidences = []
    class_ids = []

    # Check for humans in the detected objects
    for out in outs:
        for detection in out:
            scores = detection[5:]
            class_id = np.argmax(scores)
            confidence = scores[class_id]

            # If confidence is high and the class detected is a person (class_id=0 for COCO dataset)
            if confidence > 0.5 and class_id == 0:
                center_x = int(detection[0] * width)
                center_y = int(detection[1] * height)
                w = int(detection[2] * width)
                h = int(detection[3] * height)

                # Coordinates for bounding box
                x = int(center_x - w / 2)
                y = int(center_y - h / 2)

                # Add to boxes, confidences, and class_ids
                boxes.append([x, y, w, h])
                confidences.append(float(confidence))
                class_ids.append(class_id)

    # Apply Non-Maximum Suppression to remove redundant boxes
    indexes = cv2.dnn.NMSBoxes(boxes, confidences, 0.5, 0.4)

    # Only consider the first detected person (if any)
    human_detected = False  # Reset human detection flag before checking
    if len(indexes) > 0:
        for i in indexes.flatten():
            x, y, w, h = boxes[i]
            if w * h > 10000:  # Check if the detected person is big enough (close to the camera)
                human_detected = True
                person_detected_frame = frame

                # Draw bounding box and label
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                detection_label = "Employee Detected"
                break

    # If no human is detected, check the time since last detection
    if not human_detected:
            current_time = datetime.now().strftime("%H:%M:%S")
            detection_label = "Employee Not Detected"
            
            time_since_last_detection = time.time() - last_detection_time
            time_remaining = int(empthreshold - time_since_last_detection)

            if time_remaining > 0:
                countdown_text = f"Sending alert in: {time_remaining} sec"
                cv2.putText(frame, countdown_text, (20, 90), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
            else:
                if not email_sent:
                    log_status("Employee Not Detected", current_time)
                    print("Buzzer: Employee not detected ")
                    engine.say("Employee not detected")
                    engine.runAndWait()

                    image_path = "employee.jpg"
                    cv2.imwrite(image_path, frame)

                    detected_time_str = datetime.fromtimestamp(last_detection_time).strftime("%H:%M:%S")
                    subject = f"Employee Not Detected Since {detected_time_str}"
                    body = f"The employee has not been detected since {detected_time_str}.\nPlease see the attached image."
                    print("Sending email...")
                    send_email(image_path, subject, body)
                    
                    email_sent = True
                    time.sleep(5)


    if human_detected and email_sent:
        current_time = datetime.now().strftime("%H:%M:%S")
        log_status("Employee Detected", current_time)
        email_sent = False


    # Display the label on the frame
    cv2.putText(frame, detection_label, (20, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 255), 2)

    # Display the video feed
    cv2.imshow("Frame", frame)

    # Break the loop if 'q' is pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

    # Update the last detection time if a human is detected
    if human_detected:
        last_detection_time = time.time()

# Release the capture and close the windows
cap.release()
cv2.destroyAllWindows()
